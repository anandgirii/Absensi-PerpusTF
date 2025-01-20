from flask import Flask, render_template, request, session, redirect, url_for, Response, jsonify
import mysql.connector
import cv2
from PIL import Image
import numpy as np
import os
import io
import time
import xlwt
from flask import Flask, render_template, request, redirect, url_for, flash, session
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import date
from datetime import datetime
from openpyxl import Workbook
import pymysql

app = Flask(__name__, template_folder='templates', static_folder='static')

cnt = 0
pause_cnt = 0
justscanned = False

app.secret_key = 'bebasapasaja'

mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="flask_db"
)
mycursor = mydb.cursor()

UPLOAD_FOLDER = 'dataset/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< Generate dataset >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def generate_dataset(nbr):
    face_classifier = cv2.CascadeClassifier(
        "resources/haarcascade_frontalface_default.xml")

    def face_cropped(img):
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(gray, 1.3, 5)
        # scaling factor=1.3
        # Minimum neighbor = 5

        if faces is ():
            return None
        for (x, y, w, h) in faces:
            cropped_face = img[y:y + h, x:x + w]
        return cropped_face

    cap = cv2.VideoCapture(0)

    mycursor.execute("select ifnull(max(img_id), 0) from img_dataset")
    row = mycursor.fetchone()
    lastid = row[0]

    img_id = lastid
    max_imgid = img_id + 100
    count_img = 0

    while True:
        ret, img = cap.read()
        if face_cropped(img) is not None:
            count_img += 1
            img_id += 1
            face = cv2.resize(face_cropped(img), (200, 200))
            face = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)

            file_name_path = "dataset/" + nbr + "." + str(img_id) + ".jpg"
            file_name_path2 = nbr + "." + str(img_id) + ".jpg"
            cv2.imwrite(file_name_path, face)
            cv2.putText(face, str(count_img), (50, 50), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)

            mycursor.execute("""INSERT INTO img_dataset (img_id, img_person, img_path) VALUES
                                ('{}', '{}', '{}')""".format(img_id, nbr, file_name_path2))
            mydb.commit()

            frame = cv2.imencode('.jpg', face)[1].tobytes()
            yield (b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

            if cv2.waitKey(1) == 13 or int(img_id) == int(max_imgid):
                break
                cap.release()
                cv2.destroyAllWindows()


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< Train Classifier >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
@app.route('/train_classifier/<nbr>')
def train_classifier(nbr):
    dataset_dir = "dataset"

    path = [os.path.join(dataset_dir, f) for f in os.listdir(dataset_dir)]
    faces = []
    ids = []

    for image in path:
        img = Image.open(image).convert('L');
        imageNp = np.array(img, 'uint8')
        id = int(os.path.split(image)[1].split(".")[1])

        faces.append(imageNp)
        ids.append(id)
    ids = np.array(ids)

    # Train the classifier and save
    clf = cv2.face.LBPHFaceRecognizer_create()
    clf.train(faces, ids)
    clf.write("classifier.xml")

    return redirect('/petugas')


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< Face Recognition >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def face_recognition():  # Generate frame by frame from camera
    def draw_boundary(img, classifier, scaleFactor, minNeighbors, color, text, clf):
        gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        features = classifier.detectMultiScale(gray_image, scaleFactor, minNeighbors)
        coords = []

        global capture_count  # Menghitung jumlah gambar yang diambil dalam satu sesi deteksi
        global detection_complete  # Menandai apakah deteksi telah selesai
        global last_detected_id  # ID wajah yang terakhir dideteksi
        global pause_cnt
        global justscanned

        pause_cnt += 1

        for (x, y, w, h) in features:
            cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
            id, pred = clf.predict(gray_image[y:y + h, x:x + w])
            confidence = int(100 * (1 - pred / 300))
            print(f"Confidence: {confidence}")

            if confidence > 70 and not justscanned:
                capture_count += 1

                n = (100 / 20) * capture_count
                print(n)
                # w_filled = (n / 100) * w
                w_filled = (capture_count / 20) * w

                cv2.putText(img, str(int(n)) + ' %', (x + 20, y + h + 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                            (255, 255, 255), 2, cv2.LINE_AA)

                cv2.rectangle(img, (x, y + h + 40), (x + w, y + h + 50), color, 2)
                cv2.rectangle(img, (x, y + h + 40), (x + int(w_filled), y + h + 50), (255, 255, 255), cv2.FILLED)

                mycursor.execute(
                    "SELECT a.img_person, b.prs_name FROM img_dataset a "
                    "LEFT JOIN prs_mstr b ON a.img_person = b.prs_nbr WHERE img_id = %s", (id,)
                )
                row = mycursor.fetchone()
                if row:
                    pnbr = row[0]  # Nomor personel
                    pname = row[1]  # Nama personel

                    # Jika ID wajah berubah, reset capture_count dan detection_complete
                    if last_detected_id != pnbr:
                        cnt = 0
                        capture_count = 0
                        detection_complete = False
                        last_detected_id = pnbr  # Perbarui ID wajah terakhir

                    # Hitung indeks ke berapa pada hari ini hanya jika deteksi selesai
                    if not detection_complete:
                        capture_count += 1  # Tambahkan jumlah gambar yang diambil
                        print(f"Capture Count for {pname}: {capture_count}")  # Debug jumlah gambar

                        if capture_count >= 20:  # Setelah mencapai 100 gambar
                            detection_complete = True  # Tandai deteksi selesai
                            capture_count = 0  # Reset hitungan gambar

                            mycursor.execute(
                                "SELECT COUNT(*) FROM accs_hist WHERE accs_prsn = %s AND accs_date = CURDATE()", (pnbr,)
                            )
                            scan_count = mycursor.fetchone()[0] + 1  # Jumlah scan + 1
                            scan_variable = f"{pname}_{scan_count}"  # Variabel scan seperti A_1, A_2
                            print(scan_variable)  # Print variabel scan

                            # Tentukan status IN/OUT berdasarkan indeks
                            status = "IN" if scan_count % 2 != 0 else "OUT"
                            print(f"Status: {status}")

                            # Masukkan ke dalam database
                            now = datetime.now()
                            mycursor.execute(
                                "INSERT INTO accs_hist (accs_date, accs_prsn, accs_added, status) VALUES (%s, %s, %s, %s)",
                                (date.today(), pnbr, now, status)
                            )
                            mydb.commit()
                            pause_cnt = 0
                            justscanned = True



                            # Menentukan ukuran kotak berdasarkan teks
                            (text_width, text_height), _ = cv2.getTextSize(pname, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
                            top_left = (x - 15, y - 25)  # Posisi kiri atas kotak
                            bottom_right = (
                            x - 15 + text_width + 10, y - 25 + text_height + 10)  # Posisi kanan bawah kotak

                            # Kotak biru
                            cv2.rectangle(img, top_left, bottom_right, (192, 102, 34),
                                          -1)  # Kotak dengan warna biru (BGR) dan isi penuh

                            # Teks putih di atas kotak biru
                            cv2.putText(img, pname, (x - 10, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                                        (255, 255, 255), 1, cv2.LINE_AA)



            else:
                # Jika confidence terlalu rendah, reset detection_complete
                detection_complete = False
                last_detected_id = None

                if pause_cnt > 80:
                    justscanned = False

                if confidence <= 70:
                    # Menentukan ukuran kotak berdasarkan teks
                    (text_width, text_height), _ = cv2.getTextSize('UNKNOWN', cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
                    top_left = (x, y - text_height - 15)  # Posisi kiri atas kotak (disesuaikan agar lebih di atas)
                    bottom_right = (x + text_width + 10, y - 5)  # Posisi kanan bawah kotak

                    # Kotak biru
                    cv2.rectangle(img, top_left, bottom_right, (192, 102, 34),
                                  -1)  # Kotak dengan warna biru (BGR) dan isi penuh

                    # Teks putih di atas kotak biru
                    cv2.putText(img, 'UNKNOWN', (x + 5, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1,
                                cv2.LINE_AA)
                    justscanned = False

            coords = [x, y, w, h]
        return coords

    def recognize(img, clf, faceCascade):
        coords = draw_boundary(img, faceCascade, 1.1, 10, (192, 102, 34), "Face", clf)
        return img

    # Variabel global untuk menghitung pengambilan gambar dan mendeteksi selesai
    global capture_count
    global detection_complete
    global last_detected_id
    capture_count = 0  # Hitungan gambar dimulai dari 0
    detection_complete = False  # Menandai deteksi belum selesai
    last_detected_id = None  # ID wajah terakhir yang terdeteksi

    faceCascade = cv2.CascadeClassifier("../resources/haarcascade_frontalface_default.xml")
    clf = cv2.face.LBPHFaceRecognizer_create()
    clf.read("classifier.xml")

    wCam, hCam = 500, 400
    cap = cv2.VideoCapture(0)
    cap.set(3, wCam)
    cap.set(4, hCam)

    while True:
        ret, img = cap.read()
        img = recognize(img, clf, faceCascade)

        frame = cv2.imencode('.jpg', img)[1].tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')

        key = cv2.waitKey(1)
        if key == 27:  # Esc key to stop
            break


@app.route('/petugas')
def petugas():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    mycursor.execute("select prs_nbr, prs_name, prs_skill, prs_active, prs_added from prs_mstr")
    data = mycursor.fetchall()

    return render_template('index.html', data=data)


@app.route('/admin')
def admin():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    elif session['level'] != 'Admin':
        return redirect(url_for('login'))
    mycursor.execute("select id, username, email, level, date_input from tb_users")
    data = mycursor.fetchall()

    return render_template('admin.html', data=data)


@app.route('/editadmin/<_id_admin>')
def editadmin(_id_admin):
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    elif session['level'] != 'Admin':
        return redirect(url_for('login'))

    cursor = mydb.cursor()
    sql = "select id, username, email, level, date_input from tb_users where id = '{}'".format(_id_admin)

    data = (_id_admin)
    cursor.execute(sql, data)
    row = mycursor.fetchone()
    if session['level'] == 'Admin':
        return render_template('editregistrasi.html', data=row)
    else:
        return redirect(url_for('login'))


@app.route('/editadmin_submit', methods=['POST'])
def editadmin_submit():
    id = request.form['id']
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']
    level = request.form['level']

    mycursor = mydb.cursor()

    if request.form.get('password'):
        sql = "UPDATE tb_users set username= '" + username + "', email= '" + email + "', password= '" + generate_password_hash(
            password) + "', level= '" + level + "' where id = '" + id + "'"
        mycursor.execute(sql)
        flash('Registrasi Berhasil', 'success')
    else:
        sql = "UPDATE tb_users set username= '" + username + "', email= '" + email + "', level= '" + level + "' where id = '" + id + "'"
        mycursor.execute(sql)
        flash('Registrasi Berhasil', 'success')

    mydb.commit()
    return redirect('/admin')


@app.route('/deleteadmin/<_id_admin>')
def deleteadmin(_id_admin):
    if session['level'] != 'Admin':
        return redirect(url_for('login'))

    cursor = mydb.cursor()
    sql = "delete from tb_users where id = '{}'".format(_id_admin)

    mycursor.execute(sql)

    mydb.commit()
    return redirect('/admin')


@app.route('/addprsn')
def addprsn():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    mycursor.execute("select ifnull(max(prs_nbr) + 1, 101) from prs_mstr")
    row = mycursor.fetchone()
    nbr = row[0]
    # print(int(nbr))

    return render_template('addprsn.html', newnbr=int(nbr))


@app.route('/addprsn_submit', methods=['POST'])
def addprsn_submit():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    prsnbr = request.form.get('txtnbr')
    prsname = request.form.get('txtname')
    prsskill = request.form.get('txtnim')

    mycursor.execute("""INSERT INTO prs_mstr (prs_nbr, prs_name, prs_skill) VALUES
                    ('{}', '{}', '{}')""".format(prsnbr, prsname, prsskill))
    mydb.commit()

    # return redirect(url_for('home'))
    return redirect(url_for('vfdataset_page', prs=prsnbr))


@app.route('/edit/<_prs_nbr>')
def editprsn(_prs_nbr):
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    cursor = mydb.cursor()
    sql = "select * from prs_mstr where prs_nbr = '{}'".format(_prs_nbr)

    data = (_prs_nbr)
    cursor.execute(sql, data)
    row = mycursor.fetchone()

    return render_template('editprsn.html', data=row)


@app.route('/editprsn_submit', methods=['POST'])
def editprsn_submit():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    _prsnbr = request.form.get('txtnbr')
    _prsname = request.form.get('txtname')
    _prsskill = request.form.get('txtnim')

    mycursor = mydb.cursor()
    sql = "UPDATE prs_mstr set prs_name= '" + _prsname + "', prs_skill= '" + _prsskill + "' where prs_nbr = '" + _prsnbr + "'"

    mycursor.execute(sql)

    mydb.commit()

    return redirect('/petugas')


@app.route('/delete/<_prs_nbr>')
def deleteimg(_prs_nbr):
    if not 'loggedin' in session:
        return redirect(url_for('login'))

    # Hapus file berdasarkan _prs_nbr
    folder_path = "../dataset/"
    file_prefix = f"{_prs_nbr}."
    try:
        for filename in os.listdir(folder_path):
            if filename.startswith(file_prefix) and filename.endswith(".jpg"):
                os.remove(os.path.join(folder_path, filename))
    except Exception as e:
        print(f"Error saat menghapus file: {e}")

    # Hapus data dari database
    sql_queries = [
        f"DELETE FROM img_dataset WHERE img_person = '{_prs_nbr}'",
        f"DELETE FROM prs_mstr WHERE prs_nbr = '{_prs_nbr}'",
        f"DELETE FROM accs_hist WHERE accs_prsn = '{_prs_nbr}'"
    ]

    try:
        for sql in sql_queries:
            mycursor.execute(sql)
        mydb.commit()
    except Exception as e:
        print(f"Error saat menghapus data dari database: {e}")
        mydb.rollback()

    return redirect(f'/train_classifier/{_prs_nbr}')


# registrasi
@app.route('/registrasi', methods=('GET', 'POST'))
def registrasi():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    elif session['level'] != 'Admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        level = request.form['level']
        date_input = datetime.now()

        # cek username atau email
        cursor = mydb.cursor()
        cursor.execute('SELECT * FROM tb_users WHERE username=%s OR email=%s', (username, email,))
        akun = cursor.fetchone()
        if akun is None:
            cursor.execute('INSERT INTO tb_users VALUES (NULL, %s, %s, %s, %s, %s)',
                           (username, email, generate_password_hash(password), level, date_input))
            mydb.commit()
            flash('Registrasi Berhasil', 'success')
            return redirect('/admin')
        else:
            flash('Username atau email sudah ada', 'danger')
    return render_template('registrasi.html')


# login
@app.route('/login', methods=('GET', 'POST'))
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        # cek data username
        cursor = mydb.cursor()
        cursor.execute('SELECT * FROM tb_users WHERE email=%s', (email,))
        akun = cursor.fetchone()
        if akun is None:
            flash('Login Gagal, Cek Username Anda', 'danger')
        elif not check_password_hash(akun[3], password):
            flash('Login gagal, Cek Password Anda', 'danger')
        else:
            session['loggedin'] = True
            session['username'] = akun[1]
            session['level'] = akun[4]
            if session['level'] == 'Admin':
                return redirect('/admin')
            else:
                return redirect('/petugas')

    return render_template('login.html')


# logout
@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('username', None)
    session.pop('level', None)
    return redirect(url_for('login'))


@app.route('/vfdataset_page/<prs>')
def vfdataset_page(prs):
    return render_template('gendataset.html', prs=prs)


@app.route('/vidfeed_dataset/<nbr>')
def vidfeed_dataset(nbr):
    # Video streaming route. Put this in the src attribute of an img tag
    return Response(generate_dataset(nbr), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/video_feed')
def video_feed():
    # Video streaming route. Put this in the src attribute of an img tag
    return Response(face_recognition(), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/')
def fr_page():
    """Video streaming home page."""
    mycursor.execute("select a.accs_id, a.accs_prsn, b.prs_name, b.prs_skill, a.accs_added "
                     "  from accs_hist a "
                     "  left join prs_mstr b on a.accs_prsn = b.prs_nbr "
                     " where a.accs_date = curdate() "
                     " order by 1 desc")
    data = mycursor.fetchall()

    return render_template('fr_page.html', data=data)


@app.route('/countTodayScan')
def countTodayScan():
    # Hubungkan ke database
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="flask_db"
    )
    mycursor = mydb.cursor()

    # Hitung jumlah data dengan status IN pada tanggal hari ini
    mycursor.execute("SELECT COUNT(*) FROM accs_hist WHERE accs_date = CURDATE() AND status = 'IN'")
    count_in = mycursor.fetchone()[0]
    print(count_in)

    # Hitung jumlah data dengan status OUT pada tanggal hari ini
    mycursor.execute("SELECT COUNT(*) FROM accs_hist WHERE accs_date = CURDATE() AND status = 'OUT'")
    count_out = mycursor.fetchone()[0]
    print(count_out)

    # Hitung selisih jumlah IN dan OUT
    cnt = count_in - count_out

    # Kembalikan hasil sebagai JSON
    return jsonify({'rowcount': cnt})



@app.route('/loadData', methods=['GET', 'POST'])
def loadData():
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="flask_db"
    )
    mycursor = mydb.cursor()

    # Ambil data untuk ditampilkan di tabel
    mycursor.execute("select a.accs_id, a.accs_prsn, b.prs_name, a.status, date_format(a.accs_added, '%H:%i:%s') "
                     "  from accs_hist a "
                     "  left join prs_mstr b on a.accs_prsn = b.prs_nbr "
                     " where a.accs_date = curdate() "
                     " order by 1 desc")
    data = mycursor.fetchall()

    # Hitung jumlah IN
    mycursor.execute("select count(*) "
                     "  from accs_hist "
                     " where accs_date = curdate() and status = 'IN'")
    count_in = mycursor.fetchone()[0]

    # Hitung jumlah OUT
    mycursor.execute("select count(*) "
                     "  from accs_hist "
                     " where accs_date = curdate() and status = 'OUT'")
    count_out = mycursor.fetchone()[0]

    return jsonify(response=data, count_in=count_in, count_out=count_out)


@app.route('/download_excel_all')
def downloadExcelAll():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="flask_db"
    )
    cursor = mydb.cursor(buffered=True)

    sql = "select a.accs_prsn, b.prs_name,a.accs_added, a.status from accs_hist a left join prs_mstr b on a.accs_prsn = b.prs_nbr order by 1 desc"

    cursor.execute(sql)
    mydb.commit()

    output = io.BytesIO()

    wb = Workbook()
    ws = wb.active

    ws["A1"].value = "accs_prsn"
    ws["B1"].value = "prs_name"
    ws["C1"].value = "accs_added"
    ws["D1"].value = "status"

    export = cursor.fetchall()
    for x in export:
        ws.append(x)
        wb.save(output)
        output.seek(0)
    mydb.close()

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=employee_report.xlsx"})


@app.route('/download_excel_byname')
def downloadExcelByname():
    if not 'loggedin' in session:
        return redirect(url_for('login'))

    mycursor.execute("select prs_nbr, prs_name, prs_skill, prs_active, prs_added from prs_mstr")
    data = mycursor.fetchall()

    return render_template('exportexcelbyname.html', data=data)


@app.route('/export_excel_byname', methods=['POST'])
def exportExcelByname():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="flask_db"
    )

    _prsnbr = str(request.form.get('txtnbr'))
    _dateawal = str(request.form.get('txtdateawal'))
    _dateakhir = str(request.form.get('txtdateakhir'))

    # _prsnbr = '108'
    # _dateawal = '2024-07-01'
    # _dateakhir = '2024-07-31'

    mycursor = mydb.cursor(buffered=True)

    sql = "select a.accs_prsn, b.prs_name,a.accs_added, a.status from accs_hist a left join prs_mstr b on a.accs_prsn = b.prs_nbr where accs_prsn = '" + _prsnbr + "' and accs_date between '" + _dateawal + "' and '" + _dateakhir + "' order by 1 desc"
    # sql="UPDATE prs_mstr set prs_name= '" + _prsname + "', prs_skill= '" + _prsskill + "' where prs_nbr = '" + _prsnbr + "'"

    mycursor.execute(sql)

    mydb.commit()

    output = io.BytesIO()

    wb = Workbook()
    ws = wb.active

    ws["A1"].value = "accs_prsn"
    ws["B1"].value = "prs_name"
    ws["C1"].value = "accs_added"
    ws["D1"].value = "status"

    export = mycursor.fetchall()
    for x in export:
        ws.append(x)
        wb.save(output)
        output.seek(0)
    mydb.close()

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=employee_report.xlsx"})


@app.route('/download_excel_bydate')
def downloadExcelBydate():
    if not 'loggedin' in session:
        return redirect(url_for('login'))

    return render_template('exportexcelbydate.html')


@app.route('/export_excel_bydate', methods=['POST'])
def exportExcelBydate():
    if not 'loggedin' in session:
        return redirect(url_for('login'))
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="",
        database="flask_db"
    )

    _dateawal = str(request.form.get('txtdateawal'))
    _dateakhir = str(request.form.get('txtdateakhir'))

    # _prsnbr = '108'
    # _dateawal = '2024-07-01'
    # _dateakhir = '2024-07-31'

    mycursor = mydb.cursor(buffered=True)

    sql = "select a.accs_prsn, b.prs_name,a.accs_added, a.status from accs_hist a left join prs_mstr b on a.accs_prsn = b.prs_nbr where accs_date between '" + _dateawal + "' and '" + _dateakhir + "' order by 1 desc"
    # sql="UPDATE prs_mstr set prs_name= '" + _prsname + "', prs_skill= '" + _prsskill + "' where prs_nbr = '" + _prsnbr + "'"

    mycursor.execute(sql)

    mydb.commit()

    output = io.BytesIO()

    wb = Workbook()
    ws = wb.active

    ws["A1"].value = "accs_prsn"
    ws["B1"].value = "prs_name"
    ws["C1"].value = "accs_added"
    ws["D1"].value = "status"

    export = mycursor.fetchall()
    for x in export:
        ws.append(x)
        wb.save(output)
        output.seek(0)
    mydb.close()

    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=employee_report.xlsx"})

if __name__ == "__main__":
    app.run(host='127.0.0.1', port=5000, debug=True)